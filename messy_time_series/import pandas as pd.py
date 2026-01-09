import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import os as os

# 1. Setup Data Parameters
num_entries = 250
start_date = datetime(2023, 1, 1)
departments = ['Sales', 'HR', 'IT', 'Logistics', 'Marketing']
messy_values = [None, 'TBD', 'Pending', 'Error', '-']

data = []

# 2. Generate Messy Data
for i in range(num_entries):
    # Create inconsistent date formats
    current_date = start_date + timedelta(days=i)
    date_format_choice = random.choice([
        "%Y-%m-%d",           # 2023-01-01
        "%m/%d/%Y",           # 01/01/2023
        "%d-%b-%y",           # 01-Jan-23
        "%B %d, %Y",          # January 01, 2023
        "format_error"        # Occasional raw text error
    ])
    
    if date_format_choice == "format_error":
        date_str = f"Date: {current_date.month}.{current_date.day}"
    else:
        date_str = current_date.strftime(date_format_choice)

    # Messy Sales Data (Integers mixed with Strings and None)
    if random.random() < 0.1: # 10% chance of mess
        sales = random.choice(messy_values)
    else:
        sales = random.randint(1000, 50000)

    # Department (Sorted to allow for merging later)
    dept = departments[(i // 50) % 5] 

    data.append([date_str, dept, sales, f"Transaction_{i+1}"])

# Create DataFrame
df = pd.DataFrame(data, columns=['Transaction Date', 'Department', 'Amount', 'ID'])

# 3. Create Excel File with Merged Cells
wb = Workbook()
ws = wb.active
ws.title = "Messy Data"

# Write headers
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Apply Merging to 'Department' column (Column B)
# We merge every 10 rows to simulate "grouped" reporting
row_start = 2 # Skip header
while row_start < num_entries + 2:
    row_end = min(row_start + 9, num_entries + 1)
    
    # Merge cells in Column B (Department)
    ws.merge_cells(start_row=row_start, start_column=2, end_row=row_end, end_column=2)
    
    # Center align the merged cell
    cell = ws.cell(row=row_start, column=2)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_start = row_end + 1

# Save the file
wb.save("messy_data.xlsx")
print("File 'messy_data.xlsx' created successfully with 250 rows.")

# Print path where file was saved
os.path.abspath("messy_data.xlsx")
